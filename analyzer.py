import os
import zipfile
import requests
import json
import logging
from typing import List, Dict, Optional
import PyPDF2
import docx
import openpyxl
from PIL import Image
import pytesseract
from config import OPENAI_API_KEY
from document_filter import filter_documents, cleanup_temp_dirs, collect_clean_texts
from text_cleaner import preprocess_parsed_text
from prompt_builder import structured_prompt_builder

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TenderDocumentAnalyzer:
    """
    Класс для анализа документов тендера с помощью OpenAI
    """
    
    def __init__(self):
        self.openai_api_key = OPENAI_API_KEY
        self.openai_api_url = "https://api.openai.com/v1/chat/completions"
        self.supported_extensions = {
            '.doc': self._extract_text_from_doc,
            '.docx': self._extract_text_from_docx,
            '.xls': self._extract_text_from_xls,
            '.xlsx': self._extract_text_from_xlsx,
            '.pdf': self._extract_text_from_pdf,
            '.rtf': self._extract_text_from_rtf,
            '.jpg': self._extract_text_from_image,
            '.jpeg': self._extract_text_from_image,
            '.bmp': self._extract_text_from_image,
            '.png': self._extract_text_from_image
        }
    
    def extract_text_from_file(self, file_path: str) -> str:
        """
        Извлекает текст из файла в зависимости от его расширения
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            str: Извлеченный текст
        """
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext in self.supported_extensions:
                return self.supported_extensions[file_ext](file_path)
            else:
                logger.warning(f"Неподдерживаемое расширение файла: {file_ext}")
                return f"Файл {os.path.basename(file_path)} не поддерживается для анализа"
                
        except Exception as e:
            logger.error(f"Ошибка при извлечении текста из {file_path}: {e}")
            return f"Ошибка при обработке файла {os.path.basename(file_path)}: {str(e)}"
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Извлекает текст из PDF файла"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            return f"Ошибка при чтении PDF: {str(e)}"
    
    def _extract_text_from_docx(self, file_path: str) -> str:
        """Извлекает текст из DOCX файла"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            return f"Ошибка при чтении DOCX: {str(e)}"
    
    def _extract_text_from_doc(self, file_path: str) -> str:
        """Извлекает текст из DOC файла (базовая реализация)"""
        # Для .doc файлов может потребоваться дополнительная библиотека
        return f"DOC файл {os.path.basename(file_path)} требует специальной обработки"
    
    def _extract_text_from_xlsx(self, file_path: str) -> str:
        """Извлекает текст из XLSX файла"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Лист: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        except Exception as e:
            return f"Ошибка при чтении XLSX: {str(e)}"
    
    def _extract_text_from_xls(self, file_path: str) -> str:
        """Извлекает текст из XLS файла (базовая реализация)"""
        return f"XLS файл {os.path.basename(file_path)} требует специальной обработки"
    
    def _extract_text_from_rtf(self, file_path: str) -> str:
        """Извлекает текст из RTF файла"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
                # Простое удаление RTF тегов
                import re
                text = re.sub(r'\\[a-z0-9-]+\d?', '', content)
                text = re.sub(r'\{[^}]*\}', '', text)
                return text
        except Exception as e:
            return f"Ошибка при чтении RTF: {str(e)}"
    
    def _extract_text_from_image(self, file_path: str) -> str:
        """Извлекает текст из изображения с помощью OCR"""
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image, lang='rus+eng')
            return text
        except Exception as e:
            return f"Ошибка при OCR обработке изображения: {str(e)}"
    
    def analyze_tender_documents(self, file_paths: List[str], tender_info: Dict) -> Dict:
        """
        Анализирует документы тендера с помощью OpenAI
        
        Args:
            file_paths: Список путей к файлам для анализа
            tender_info: Информация о тендере
            
        Returns:
            Dict: Результат анализа
        """
        try:
            # Используем новый пайплайн collect_clean_texts
            logger.info(f"Запускаю полный пайплайн обработки для {len(file_paths)} файлов")
            
            # Получаем очищенные тексты через collect_clean_texts
            # Извлекаем номер тендера из tender_info
            tender_number = tender_info.get('TenderNumOuter', 'unknown')
            clean_result = collect_clean_texts(file_paths, tender_number)
            
            if not clean_result['success']:
                return {
                    'success': False,
                    'error': clean_result.get('error', 'Ошибка при обработке документов')
                }
            
            clean_text = clean_result['text']
            items = clean_result.get('items', [])
            
            # Подготавливаем данные для structured_prompt_builder
            tender_data = {
                "reg_number": tender_info.get('TenderNumOuter', 'Не указан'),
                "title": tender_info.get('TenderName', 'Не указано'),
                "customer": tender_info.get('Customer', 'Не указан'),
                "region": tender_info.get('Region', 'Не указан'),
                "price": f"{tender_info.get('Price', '0')} ₽",
                "deadline": tender_info.get('EndTime', 'Не указан'),
                "tender_url": tender_info.get('TenderLink', ''),
                "tenderguru_url": tender_info.get('TenderLinkInner', '')
            }
            
            # Формируем структурированный промпт
            prompt = structured_prompt_builder(
                tender_data=tender_data,
                clean_text=clean_text,
                items=items
            )
            
            # Отправляем запрос к OpenAI
            analysis_result = self._send_to_openai(prompt)
            
            # Очищаем временные папки после завершения анализа
            cleanup_temp_dirs()
            
            return {
                'success': True,
                'analysis': analysis_result,
                'files_processed': clean_result.get('files_processed', 0),
                'total_files': len(file_paths),
                'filtered_files': clean_result.get('filtered_files', []),
                'clean_text': clean_text,
                'items': items,
                'stats': clean_result.get('stats', {})
            }
            
        except Exception as e:
            logger.error(f"Ошибка при анализе документов: {e}")
            # Очищаем временные папки даже в случае ошибки
            cleanup_temp_dirs()
            return {
                'success': False,
                'error': str(e)
            }
    

    
    def _send_to_openai(self, prompt: str) -> str:
        """
        Отправляет запрос к OpenAI API
        
        Args:
            prompt: Промпт для анализа
            
        Returns:
            str: Ответ от OpenAI
        """
        try:
            headers = {
                'Authorization': f'Bearer {self.openai_api_key}',
                'Content-Type': 'application/json'
            }
            
            data = {
                'model': 'gpt-3.5-turbo',
                'messages': [
                    {
                        'role': 'system',
                        'content': 'Ты эксперт по анализу тендеров и государственных закупок. Твоя задача - предоставить детальный анализ документов тендера и дать практические рекомендации.'
                    },
                    {
                        'role': 'user',
                        'content': prompt
                    }
                ],
                'max_tokens': 3000,
                'temperature': 0.7
            }
            
            response = requests.post(self.openai_api_url, headers=headers, json=data)
            
            if response.status_code == 200:
                result = response.json()
                return result['choices'][0]['message']['content']
            else:
                logger.error(f"Ошибка OpenAI API: {response.status_code} - {response.text}")
                return f"Ошибка при обращении к OpenAI API: {response.status_code}"
                
        except Exception as e:
            logger.error(f"Ошибка при отправке запроса к OpenAI: {e}")
            return f"Ошибка при анализе: {str(e)}"

# Функция для удобного использования
def analyze_tender_documents(file_paths: List[str], tender_info: Dict) -> Dict:
    """
    Удобная функция для анализа документов тендера
    
    Args:
        file_paths: Список путей к файлам
        tender_info: Информация о тендере
        
    Returns:
        Dict: Результат анализа
    """
    analyzer = TenderDocumentAnalyzer()
    return analyzer.analyze_tender_documents(file_paths, tender_info)

 